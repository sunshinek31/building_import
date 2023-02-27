import openpyxl
import shutil
from openpyxl import load_workbook
import glob, os

# 循环读取excel,记录正在读取操作的excel的文件名
# 循环读取excel的sheet,记录正在读取操作的sheet名
# 循环读取sheet的每行数据,记录正在读取操作sheet的行数

#讲读取到的数据, 写入到temp文件中
#核心数据: 楼宇名、房间号、用房责任人、二级使用单位
#辅助数据: 数据对应的excel文件名,数据对应的sheet名,数据对应的sheet行数

tempFolder = '../temp'
tempImportDataExcelPath = '../temp/tempImportData.xlsx'
tempErrorImportDataCsvPath = '../temp/tempErrorImportData.csv'

def load_all_excel_import_data(*read_line_operation):
    excel_data_folder = "../db/"
    os.chdir(excel_data_folder)
    for file in glob.glob("*.xlsx"):
        if not file.startswith("~$"):
            # 取得合法文件, 开始读取内容
            read_excel_about_import_data(file)

def read_excel_about_import_data(excel_file):
    wb = load_workbook(excel_file)
    print(wb.sheetnames)

    building_name_col = 2
    building_room_name_col = 3
    username_col = 8
    second_department_col = 10

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print('当前sheet(' + sheet.title + ')的拥有' + str(sheet.max_row - 1) + '行待处理数据')

        for row in range(2,sheet.max_row+1):
            if sheet.max_column < 10 :
                with open(tempErrorImportDataCsvPath,'a') as tempErrorImportDataCsvFile:
                    print(excel_file.title()+','+sheetname+','+str(row),file=tempErrorImportDataCsvFile)
                    # tempErrorImportDataCsvFile.write(excel_file.title()+','+sheetname+','+str(row)+'\n')
            else:

                building_name = sheet.cell(row=row,column=building_name_col).value
                building_room_name = sheet.cell(row=row,column=building_room_name_col).value
                username = sheet.cell(row=row,column=username_col).value
                second_department = sheet.cell(row=row,column=second_department_col).value
                print(building_name, building_room_name, username, second_department, excel_file.title(), sheetname,
                      row)

                # 存在没有责任人的数据, 直接判断异常并暂存入文件
                if username is None :
                    with open(tempErrorImportDataCsvPath, 'a') as tempErrorImportDataCsvFile:
                        print(excel_file.title() + ',' + sheetname + ',' + str(row), file=tempErrorImportDataCsvFile)
                else:
                    #加载
                    tempImportDataExcel = load_workbook(tempImportDataExcelPath)
                    tempDataSheet = tempImportDataExcel['Sheet']
                    tempDataSheet.append([building_name, building_room_name, username, second_department, excel_file.title(), sheetname,
                      row])
                    tempImportDataExcel.save(tempImportDataExcelPath)



if os.path.exists(tempFolder):
    print("存在temp文件夹")
    print("删除文件夹...")

    shutil.rmtree(tempFolder)

os.makedirs(tempFolder, 0o755)
#创建零时excel,存放核心数据
wb = openpyxl.Workbook()
wb.save(tempImportDataExcelPath)

load_all_excel_import_data()


