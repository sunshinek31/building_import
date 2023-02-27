from openpyxl import load_workbook
import glob, os


def read_import_data_excel(excel_file):
    wb = load_workbook(excel_file)
    print(wb.sheetnames)


os.chdir('../db/')
for file in glob.glob("*.xlsx"):
    if not file.startswith("~$"):
        print(file)
        read_import_data_excel(file)
